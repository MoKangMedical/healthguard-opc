"""
数据导出服务
"""

from typing import Dict, Any, List, Optional
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
import csv
import io
import json
from abc import ABC, abstractmethod

from app.models.patient import Patient
from app.models.health_record import (
    HealthRecord, BloodPressure, BloodSugar, HeartRate, Weight, HealthRecordType
)
from app.models.appointment import Appointment
from app.models.medication import Medication
from app.models.device import Device, DeviceDataRecord

class BaseExporter(ABC):
    """导出器基类"""
    
    @abstractmethod
    def export(self, data: List[Dict], headers: List[str]) -> bytes:
        pass

class CSVExporter(BaseExporter):
    """CSV 导出器"""
    
    def export(self, data: List[Dict], headers: List[str]) -> bytes:
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue().encode('utf-8-sig')

class JSONExporter(BaseExporter):
    """JSON 导出器"""
    
    def export(self, data: List[Dict], headers: List[str]) -> bytes:
        return json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')

class ExcelExporter(BaseExporter):
    """Excel 导出器"""
    
    def export(self, data: List[Dict], headers: List[str]) -> bytes:
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(header, ''))
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        except ImportError:
            # 如果没有 openpyxl，回退到 CSV
            return CSVExporter().export(data, headers)


class ExportService:
    """数据导出服务"""
    
    def __init__(self, db: Session):
        self.db = db
        self.exporters = {
            'csv': CSVExporter(),
            'json': JSONExporter(),
            'xlsx': ExcelExporter(),
        }
    
    def export_health_records(
        self,
        patient_id: int,
        start_date: datetime,
        end_date: datetime,
        format: str = 'csv'
    ) -> Dict[str, Any]:
        """导出健康记录"""
        
        records = self.db.query(HealthRecord).filter(
            HealthRecord.patient_id == patient_id,
            HealthRecord.record_date >= start_date,
            HealthRecord.record_date <= end_date
        ).order_by(HealthRecord.record_date.desc()).all()
        
        data = []
        for record in records:
            row = {
                '日期': record.record_date.strftime('%Y-%m-%d %H:%M'),
                '类型': record.record_type.value,
                '来源': record.source or '手动',
                '异常': '是' if record.is_abnormal else '否',
            }
            
            # 添加具体数据
            if record.blood_pressure:
                row['收缩压'] = record.blood_pressure.systolic
                row['舒张压'] = record.blood_pressure.diastolic
                row['脉搏'] = record.blood_pressure.pulse
            elif record.blood_sugar:
                row['血糖值'] = record.blood_sugar.value
                row['单位'] = record.blood_sugar.unit
                row['测量时间'] = record.blood_sugar.measurement_time
            elif record.heart_rate:
                row['心率'] = record.heart_rate.bpm
                row['静息状态'] = '是' if record.heart_rate.is_resting else '否'
            elif record.weight:
                row['体重'] = record.weight.value
                row['BMI'] = record.weight.bmi
            
            data.append(row)
        
        headers = list(set().union(*(d.keys() for d in data))) if data else []
        
        exporter = self.exporters.get(format, CSVExporter())
        content = exporter.export(data, headers)
        
        return {
            'content': content,
            'filename': f'health_records_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.{format}',
            'content_type': self._get_content_type(format)
        }
    
    def export_medications(
        self,
        patient_id: int,
        format: str = 'csv'
    ) -> Dict[str, Any]:
        """导出用药记录"""
        
        medications = self.db.query(Medication).filter(
            Medication.patient_id == patient_id
        ).order_by(Medication.created_at.desc()).all()
        
        data = [
            {
                '药物名称': m.medication_name,
                '通用名': m.generic_name or '',
                '剂量': m.dosage or '',
                '频率': m.frequency or '',
                '用药途径': m.route or '',
                '开始日期': m.start_date.strftime('%Y-%m-%d') if m.start_date else '',
                '结束日期': m.end_date.strftime('%Y-%m-%d') if m.end_date else '',
                '总量': m.quantity or '',
                '剩余量': m.remaining or '',
                '状态': '活跃' if m.is_active else '已停用',
                '说明': m.instructions or '',
            }
            for m in medications
        ]
        
        headers = ['药物名称', '通用名', '剂量', '频率', '用药途径', '开始日期', '结束日期', '总量', '剩余量', '状态', '说明']
        
        exporter = self.exporters.get(format, CSVExporter())
        content = exporter.export(data, headers)
        
        return {
            'content': content,
            'filename': f'medications_{datetime.now().strftime("%Y%m%d")}.{format}',
            'content_type': self._get_content_type(format)
        }
    
    def export_appointments(
        self,
        patient_id: int,
        start_date: datetime,
        end_date: datetime,
        format: str = 'csv'
    ) -> Dict[str, Any]:
        """导出预约记录"""
        
        appointments = self.db.query(Appointment).filter(
            Appointment.patient_id == patient_id,
            Appointment.appointment_date >= start_date,
            Appointment.appointment_date <= end_date
        ).order_by(Appointment.appointment_date.desc()).all()
        
        data = [
            {
                '预约号': a.appointment_no or '',
                '预约日期': a.appointment_date.strftime('%Y-%m-%d %H:%M'),
                '科室': a.department or '',
                '类型': a.appointment_type or '',
                '医生': a.doctor.full_name if a.doctor else '',
                '状态': a.status.value,
                '原因': a.reason or '',
            }
            for a in appointments
        ]
        
        headers = ['预约号', '预约日期', '科室', '类型', '医生', '状态', '原因']
        
        exporter = self.exporters.get(format, CSVExporter())
        content = exporter.export(data, headers)
        
        return {
            'content': content,
            'filename': f'appointments_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.{format}',
            'content_type': self._get_content_type(format)
        }
    
    def export_device_data(
        self,
        device_id: int,
        start_date: datetime,
        end_date: datetime,
        format: str = 'csv'
    ) -> Dict[str, Any]:
        """导出设备数据"""
        
        records = self.db.query(DeviceDataRecord).filter(
            DeviceDataRecord.device_id == device_id,
            DeviceDataRecord.measured_at >= start_date,
            DeviceDataRecord.measured_at <= end_date
        ).order_by(DeviceDataRecord.measured_at.desc()).all()
        
        data = [
            {
                '测量时间': r.measured_at.strftime('%Y-%m-%d %H:%M:%S'),
                '数据类型': r.data_type or '',
                '数据质量': r.quality or '',
                '解析数据': json.dumps(r.parsed_data, ensure_ascii=False) if r.parsed_data else '',
            }
            for r in records
        ]
        
        headers = ['测量时间', '数据类型', '数据质量', '解析数据']
        
        exporter = self.exporters.get(format, CSVExporter())
        content = exporter.export(data, headers)
        
        return {
            'content': content,
            'filename': f'device_data_{device_id}_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.{format}',
            'content_type': self._get_content_type(format)
        }
    
    def export_patient_summary(
        self,
        patient_id: int,
        format: str = 'csv'
    ) -> Dict[str, Any]:
        """导出患者摘要"""
        
        patient = self.db.query(Patient).filter(Patient.id == patient_id).first()
        if not patient:
            raise ValueError("患者不存在")
        
        data = [{
            '病历号': patient.patient_no or '',
            '姓名': patient.user.full_name if patient.user else '',
            '性别': patient.gender or '',
            '年龄': patient.age or '',
            '血型': patient.blood_type or '',
            '地址': patient.address or '',
            '紧急联系人': patient.emergency_contact or '',
            '紧急联系电话': patient.emergency_phone or '',
            '过敏史': patient.allergies or '',
            '病史': patient.medical_history or '',
            '糖尿病': '是' if patient.has_diabetes else '否',
            '高血压': '是' if patient.has_hypertension else '否',
            '心脏病': '是' if patient.has_heart_disease else '否',
        }]
        
        headers = list(data[0].keys())
        
        exporter = self.exporters.get(format, CSVExporter())
        content = exporter.export(data, headers)
        
        return {
            'content': content,
            'filename': f'patient_{patient.patient_no}_{datetime.now().strftime("%Y%m%d")}.{format}',
            'content_type': self._get_content_type(format)
        }
    
    def _get_content_type(self, format: str) -> str:
        """获取内容类型"""
        types = {
            'csv': 'text/csv',
            'json': 'application/json',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        return types.get(format, 'application/octet-stream')